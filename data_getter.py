from pathlib import Path
from openpyxl import load_workbook


class DataGetter:
    def __init__(self, xlsx_path: str):
        '''
        Clase para obtener datos de un archivo Excel.
        Contiene métodos que implementan la lógica para leer nuestros datos.
        '''
        self.xlsx_path = xlsx_path
        self.wb = load_workbook(self.xlsx_path, data_only=True)

    def get_column_values(self, sheet_name: str, col: str, start_row: int = 2):
        '''
        Obtiene los valores de una columna específica a partir de una fila dada.
        Asume que la primera fila contiene encabezados y que los datos comienzan desde la fila 2 por defecto.
        '''
        try:
            ws = self.wb[sheet_name]
            return [
                ws[f"{col}{row}"].value
                for row in range(start_row, ws.max_row + 1)
                if ws[f"{col}{row}"].value is not None
            ]
        except KeyError:
            print(f"Error: La hoja '{sheet_name}' no existe en el archivo '{self.xlsx_path}'.")
            return []
        except Exception as e:
            print(f"Error al leer la columna '{col}' en la hoja '{sheet_name}': {e}")
            return []

    def get_table_values(self, sheet_name: str, col: str, *parameters):
        '''
        Obtiene los valores de una columna específica, la cual es luego dividida en segmentos según los parámetros dados.
        Devuelve una lista de listas (filas) con los valores correspondientes.
        '''
        try:
            ws = self.wb[sheet_name]
            start_row: int = 2

            vals = [
                ws[f"{col}{row}"].value
                for row in range(start_row, ws.max_row + 1)
                if ws[f"{col}{row}"].value is not None
            ]
            if not parameters:
                return vals

            dims = [len(p) for p in parameters]

            def nest(flat_list, shape):
                if not shape:
                    return flat_list
                dim = shape[0]
                chunk_size = -(-len(flat_list) // dim)  # ceil division
                return [nest(flat_list[i * chunk_size:(i + 1) * chunk_size], shape[1:]) for i in range(dim)]

            return nest(vals, dims)
        except KeyError:
            print(f"Error: La hoja '{sheet_name}' no existe en el archivo '{self.xlsx_path}'.")
            return []
        except Exception as e:
            print(f"Error al leer la tabla en la hoja '{sheet_name}': {e}")
            return []